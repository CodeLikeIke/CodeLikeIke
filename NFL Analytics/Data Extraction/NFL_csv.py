import csv 
import os
from openpyxl import load_workbook

wb = load_workbook(filename="AFCstandings.xlsx")
wb = load_workbook(filename="NFCstandings.xlsx")
wb = load_workbook(filename="AFCseed.xlsx")
wb = load_workbook(filename="NFCseed.xlsx")
wb = load_workbook(filename="results.xlsx")


sheet = wb.active

csv_data=[]
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))
    
with open('AFCstandings.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
    for row in csv_data:
        print(row)

with open('NFCstandings.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
    for row in csv_data:
        print(row)     
        
with open('AFCseed.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
    for row in csv_data:
        print(row) 
        
with open('NFCseed.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
    for row in csv_data:
        print(row)        
        
with open('results.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
    for row in csv_data:
        print(row)
        
                   
     
        